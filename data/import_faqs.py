"""
Importa data/faqs.xlsx a la colección `faq` en MongoDB.

Uso:
    python data/import_faqs.py            # inserta solo FAQs nuevas (por pregunta)
    python data/import_faqs.py --limpiar  # vacía la colección antes de insertar
"""
import sys
import os
import datetime

# ── Ajusta el path para que database.py sea importable desde aquí ──
project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if project_root not in sys.path:
    sys.path.insert(0, project_root)

import openpyxl
from database import faq_collection

EXCEL_PATH = os.path.join(os.path.dirname(__file__), "faqs.xlsx")


def cargar_excel():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1] and row[2]:   # pregunta y respuesta no vacías
            rows.append({"pregunta": str(row[1]).strip(), "respuesta": str(row[2]).strip()})
    return rows


def importar(limpiar: bool):
    faqs = cargar_excel()
    print(f"📄 FAQs leídas del Excel: {len(faqs)}")

    if limpiar:
        eliminados = faq_collection.delete_many({}).deleted_count
        print(f"🗑️  Colección limpiada ({eliminados} documentos eliminados).")
        existentes = set()
    else:
        existentes = {doc["pregunta"] for doc in faq_collection.find({}, {"pregunta": 1, "_id": 0})}
        print(f"ℹ️  FAQs ya en MongoDB: {len(existentes)}")

    ahora = datetime.datetime.utcnow()
    nuevas = []
    for faq in faqs:
        if faq["pregunta"] not in existentes:
            nuevas.append({
                "pregunta":       faq["pregunta"],
                "respuesta":      faq["respuesta"],
                "bloqueado":      True,
                "categoria":      "general",
                "fuente":         "faqs.xlsx",
                "fecha_creacion": ahora,
            })

    if not nuevas:
        print("✅ No hay FAQs nuevas para insertar.")
        return

    faq_collection.insert_many(nuevas)
    print(f"✅ Insertadas: {len(nuevas)} FAQs nuevas.")
    print(f"📊 Total en colección ahora: {faq_collection.count_documents({})}")


if __name__ == "__main__":
    limpiar = "--limpiar" in sys.argv
    if limpiar:
        print("⚠️  Modo --limpiar: se borrará toda la colección faq antes de importar.")
        confirmacion = input("¿Continuar? (s/N): ").strip().lower()
        if confirmacion != "s":
            print("Cancelado.")
            sys.exit(0)
    importar(limpiar)
